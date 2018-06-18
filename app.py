import openpyxl
import PIL
from PIL import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import sys
import os
# declare acct details
accounts = {
            1:{'name': 'Clover Ridge Plaza', 'monthly': 2120.00, 'address': '255 Lesmill Road', 'postal': 'Don Mills  ON  M3B 2V1', 'num': '06'},
            2:{'name': 'Rosslyn Plaza', 'monthly': 815.00, 'address': '255 Lesmill Road', 'postal': 'Don Mills  ON  M3B 2V1', 'num': '05'},
            3:{'name': 'ATX Network Corp.', 'monthly': 710.00, 'address': '501 Clements West, Unit#1', 'postal': 'Ajax  ON  L1S 7H4', 'num': '09'},
            4:{'name': 'ATX Network Corp. (Green St.)', 'monthly': 570.00, 'address': '501 Clements West, Unit#1', 'postal': 'Ajax  ON  L1S 7H4', 'num': '10'},
            5:{'name': 'Ajax Tire', 'monthly': 400.00, 'address': '369 Finley Avenue', 'postal': 'Ajax  ON  L1S 2E2', 'num': '14'},
            6:{'name': 'Dwight Crane Ltd.', 'monthly': 495.00, 'address': '131 Dowty Road', 'postal': 'Ajax  ON  L1S 2G3', 'num': '18'},
            7:{'name': 'RTS Packaging Canada Inc.', 'monthly': 400.00, 'address': '782 McKay Road', 'postal': 'Pickering  ON  L1W 2Y4', 'num': '17'},
            8:{'name': 'Durham Condo', 'monthly': 450.00, 'address': '218 Dundas St. East', 'postal': 'Whitby  ON  L1N 2H8', 'num': '12'},
            9:{'name': 'Index Energy', 'monthly': 545.00, 'address': '170 Mills Road', 'postal': 'Ajax  ON  L1S 2H1', 'num': '25'},
            }

def save_file(date):
# generate all with no changes
    for account in accounts:
      os.chdir('C:/Users/Nick/Desktop/invoice_test')
      wb = openpyxl.load_workbook('C:/Users/Nick/Documents/CODE/INVOICE GEN/TEMPLATE.xlsx')
      ws = wb['Service Invoice']
      header = Image('invoice_head.png')
      ws.add_image(header, 'A1')
      ws['A13'].value = accounts[account]['name']
      ws['A14'].value = accounts[account]['address']
      ws['A15'].value = accounts[account]['postal']
      ws['A24'].value = 'Summer Grounds Maintenance - ' + date[5:]
      ws['D24'].value = accounts[account]['monthly']
      ws['D4'].value = date
      ws['D5'].value = accounts[account]['num'] + '-' + date[5:7] + date[8:] + '-00' 

      wb.save(accounts[account]['num'] + '-' + date[5:7] + date[8:] + '-00.xlsx')
      print('Generated invoice: ' + accounts[account]['name'] + ' ' + date + '.xlsx')
      result = 'Invoices generated and saved.'
    return result

def change_existing(acct_num, date, service_cost, service_desc):
  
    os.chdir('C:/Users/Nick/Desktop/invoice_test')
    wb = openpyxl.load_workbook('C:/Users/Nick/Documents/CODE/INVOICE GEN/TEMPLATE.xlsx')
    ws = wb['Service Invoice']
    header = Image('invoice_head.png')
    ws.add_image(header, 'A1')
    ws['A13'].value = accounts[acct_num]['name']
    ws['A14'].value = accounts[acct_num]['address']
    ws['A15'].value = accounts[acct_num]['postal']
    ws['A24'].value = 'Summer Grounds Maintenance - ' + date[5:]
    ws['D24'].value = accounts[acct_num]['monthly']
    ws['D4'].value = date
    ws['D5'].value = accounts[acct_num]['num'] + '-' + date[5:7] + date[8:] + '-00' 
    ws['A26'].value = service_desc
    ws['D26'].value = service_cost

    wb.save(accounts[acct_num]['num'] + '-' + date[5:7] + date[8:] + '-00.xlsx')
    print('Generated invoice: ' + accounts[acct_num]['name'] + ' ' + date + '.xlsx')
    result = 'Invoice generated and saved.'
    
    return result

def create_new(acct_num, date, service_cost, service_desc):

    os.chdir('C:/Users/Nick/Desktop/invoice_test')
    wb = openpyxl.load_workbook('C:/Users/Nick/Documents/CODE/INVOICE GEN/TEMPLATE.xlsx')
    ws = wb['Service Invoice']
    header = Image('invoice_head.png')
    ws.add_image(header, 'A1')
    ws['A13'].value = accounts[acct_num]['name']
    ws['A14'].value = accounts[acct_num]['address']
    ws['A15'].value = accounts[acct_num]['postal']
    ws['A24'].value = 'Summer Grounds Maintenance - ' + date[5:]
    ws['D24'].value = accounts[acct_num]['monthly']
    ws['D4'].value = date
    ws['D5'].value = accounts[acct_num]['num'] + '-' + date[5:7] + date[8:] + '-01' 
    ws['A24'].value = service_desc
    ws['D24'].value = service_cost

    wb.save(accounts[acct_num]['num'] + '-' + date[5:7] + date[8:] + '-01.xlsx')
    print('Generated invoice: ' + accounts[acct_num]['name'] + ' ' + date + '.xlsx')
    result = 'Invoice generated and saved.'

    return result

def main():
  while True:
    # check for other changes
    print('\nWhat would you like to do: \n1) Generate and print all invoices.\n2) Add a surcharge to an existing invoice.\n3) Create a new, seperate invoice.\n4) Exit.')
    response = int(input())

    # if no changes to existing invoice, generate and save all invoices.
    if response == 1:
      print('\nExcellent. Lets move on.')
      result = save_file(date)
      print(result)
      continue
    # if yes, which account
    elif response == 2:
      
      print('Enter the corresponding account number for the invoice you wish to modify:')

      for k in accounts:
        print(k, accounts[k]['name'])

      acct_num = int(input('ACCOUNT NUMBER: '))

    # verify selection
      print('\nYou need to make a change to the invoice for ' + accounts[acct_num]['name'] + '. Is that correct? (y/n)')
      response = input()

    # if correct selection, ask for description of service added
      if response == 'y':
        print('\nEnter the description of service added:')

        service_desc = input("Description: ")

        print('\nEnter the cost of service added (no $/no decimals):')

        service_cost = int(input("Cost: "))

        print('So we are adding a fee of $' + str(service_cost) + ' for ' + service_desc + ' at ' + accounts[acct_num]['name'] + '. Is that correct? (y/n)')

        result = change_existing(acct_num, date, service_cost, service_desc)
        print(result)
        continue
    # generate new invoice
    elif response == 3:
      
      print('Enter the corresponding account number for the invoice you wish to create:')

      for k in accounts:
        print(k, accounts[k]['name'])

      acct_num = int(input('ACCOUNT NUMBER: '))

    # verify selection
      print('\nYou need to create an invoice for ' + accounts[acct_num]['name'] + '. Is that correct? (y/n)')
      response = input()

    # if correct selection, ask for description of service added
      if response == 'y':
        print('\nEnter the description of service:')

        service_desc = input("Description: ")

        print('\nEnter the cost of service (no $ / no decimals):')

        service_cost = int(input("Cost: "))

        print('So we are creating an invoice for $' + str(service_cost) + ' for ' + service_desc + ' at ' + accounts[acct_num]['name'] + '. Is that correct? (y/n)')

        result = create_new(acct_num, date, service_cost, service_desc)
        print(result)
        continue
    elif response == 4:
      sys.exit("Goodbye.")
    # bad input
    else:
      print('Respond with 1, 2, 3 or 4 jackass.')


# Get date
print('Welcome to the APM Invoice Generator.\nPlease enter the date for the invoices (YYYY-MM-DD).')
date = input('Date:')

main()

